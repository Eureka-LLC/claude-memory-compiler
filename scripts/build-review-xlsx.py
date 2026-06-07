#!/usr/bin/env python3
"""Render review.xlsx from review-data.json for the manual legacy review.

One row per article. Columns: file/title/type/summary (reference), scope (dropdown),
project (type the real folder name), then one checkbox column per domain ("x" where
the LLM suggested it). Edit in Excel, then read-review-xlsx.py turns it back into JSON.
"""
import json
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ModuleNotFoundError:
    raise SystemExit("Нужен пакет openpyxl. Установи: py -m pip install openpyxl")

if len(sys.argv) < 2:
    raise SystemExit("Использование: build-review-xlsx.py <путь к каталогу .claude>")
BRAIN = Path(sys.argv[1])
data = json.loads((BRAIN / "review-data.json").read_text(encoding="utf-8"))
vocab = list(data["vocab"])
records = data["records"]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "review"

base = ["file", "title", "type", "scope", "project", "summary", "current"]
headers = base + vocab
ws.append(headers)

head_fill = PatternFill("solid", fgColor="DDDDDD")
for c, _ in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.fill = head_fill

for rec in records:
    doms = {str(d).lower() for d in (rec.get("domains") or [])}
    row = [rec.get("file", ""), rec.get("title", ""), rec.get("type", ""),
           rec.get("scope", ""), rec.get("project", ""), rec.get("summary", ""),
           ", ".join(rec.get("current") or [])]
    row += ["x" if v.lower() in doms else "" for v in vocab]
    ws.append(row)

nrows = len(records) + 1

# scope dropdown (column D)
dv = DataValidation(type="list", formula1='"global,project"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"D2:D{nrows}")

ws.freeze_panes = "B2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{nrows}"

widths = {"A": 34, "B": 40, "C": 9, "D": 10, "E": 20, "F": 48, "G": 20}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
for i in range(len(base) + 1, len(headers) + 1):
    ws.column_dimensions[get_column_letter(i)].width = 13
    ws.cell(row=1, column=i).alignment = Alignment(horizontal="center", text_rotation=45, wrap_text=True)

# center the domain 'x' cells
for r in range(2, nrows + 1):
    for i in range(len(base) + 1, len(headers) + 1):
        ws.cell(row=r, column=i).alignment = Alignment(horizontal="center")

out = BRAIN / "review.xlsx"
try:
    wb.save(out)
except PermissionError:
    raise SystemExit("ОШИБКА: review.xlsx занят (открыт в Excel?). Закрой файл и повтори.")
print(f"review.xlsx: {out}  (rows: {len(records)}, domains: {len(vocab)})")
