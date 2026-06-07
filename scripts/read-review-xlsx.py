#!/usr/bin/env python3
"""Read the edited review.xlsx back into review-applied.json for apply-review.ps1.

Output: [ { "file", "scope", "project", "domains": [...] }, ... ]
A domain column counts as checked if the cell contains x / х / ✓ / 1 / yes / да.
"""
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ModuleNotFoundError:
    raise SystemExit("Нужен пакет openpyxl. Установи: py -m pip install openpyxl")

if len(sys.argv) < 2:
    raise SystemExit("Использование: read-review-xlsx.py <путь к каталогу .claude>")
BRAIN = Path(sys.argv[1])
CHECKED = {"x", "х", "✓", "v", "1", "true", "y", "yes", "да", "+"}

wb = openpyxl.load_workbook(BRAIN / "review.xlsx")
ws = wb.active
hdr = [c.value for c in ws[1]]
idx = {h: i for i, h in enumerate(hdr)}
if "file" not in idx:
    raise SystemExit("ОШИБКА: в заголовке review.xlsx нет колонки 'file'.")
base_set = {"file", "title", "type", "scope", "project", "summary", "current"}
domain_cols = [h for h in hdr if h is not None and h not in base_set]

out = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or not row[idx["file"]]:
        continue
    doms = [h for h in domain_cols if str(row[idx[h]] or "").strip().lower() in CHECKED]
    out.append({
        "file":    str(row[idx["file"]]).strip(),
        "scope":   str(row[idx["scope"]] or "").strip(),
        "project": str(row[idx["project"]] or "").strip(),
        "domains": doms,
    })

dst = BRAIN / "review-applied.json"
dst.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"review-applied.json: {len(out)} записей -> {dst}")
